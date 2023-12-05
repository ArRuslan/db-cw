from datetime import timedelta, datetime
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, HTTPException, Response
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, Fill, PatternFill, Color
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pytz import UTC
from tortoise.functions import Count, Sum

from app.models import Product, Customer, Order, Category, Return
from app.models.order import OrderPd
from app.models.order_item import OrderItemPd
from app.models.product import ProductPd
from app.utils import AuthManagerDep, Permissions

router = APIRouter(prefix="/api/v0/reports")


class CustomWorksheet:
    def __init__(self, ws: Worksheet):
        self.ws = ws

    def fix_formatting(self) -> Worksheet:
        for col in self.ws.columns:
            max_len = 0
            let = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass

            self.ws.column_dimensions[let].width = (max_len+2)

        return self.ws

    def append(self, iterable: list | tuple | range | dict, **cell_kwargs):
        def update_or_create_cell(content_: Cell | str | float | datetime | int, col_idx_, row_idx_):
            if not isinstance(content_, Cell):
                return Cell(self.ws, row=row_idx, column=col_idx, value=content_, **cell_kwargs)

            cell_ = content_
            cell_.parent = self.ws
            cell_.column = col_idx_
            cell_.row = row_idx_

            return cell_

        row_idx = self.ws._current_row + 1

        if isinstance(iterable, (list, tuple, range)):
            for col_idx, content in enumerate(iterable, 1):
                self.ws._cells[(row_idx, col_idx)] = update_or_create_cell(content, col_idx, row_idx)

        elif isinstance(iterable, dict):
            for col_idx, content in iterable.items():
                if isinstance(col_idx, str):
                    col_idx = column_index_from_string(col_idx)
                self.ws._cells[(row_idx, col_idx)] = update_or_create_cell(content, col_idx, row_idx)

        else:
            self.ws._invalid_row(iterable)

        self.ws._current_row = row_idx

    def __call__(self, value: str | float | int | datetime) -> Cell:
        return Cell(self.ws, value=value)


def upd_cell(cell: Cell | str | float | int | datetime, **cell_attributes) -> Cell:
    if not isinstance(cell, Cell):
        cell = Cell(None, value=cell)

    for attr in cell_attributes:
        setattr(cell, attr, cell_attributes[attr])
    return cell


def mkfill(rgb: str, tint: float) -> Fill:
    return PatternFill(start_color=Color(rgb, tint=tint), fill_type="solid")


@router.get("/products/{product_id}")
async def product_report(manager: AuthManagerDep, product_id: int, fmt: Literal["json", "excel"]="json"):
    if not Permissions.check(manager, Permissions.READ_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient privileges!")

    product = await Product.get_or_none(id=product_id).select_related("category") \
        .annotate(order_count=Count("orderitems"), orderitem_count=Sum("orderitems__quantity"))

    result = {
        "id": product.id,
        "model": product.model,
        "manufacturer": product.manufacturer,
        "price": product.price,
        "quantity": product.quantity,
        "per_order_limit": product.per_order_limit,
        "image_url": product.image_url,
        "warranty_days": product.warranty_days,
        "category_name": product.category.name,
        "order_count": product.order_count,
        "order_item_count": product.orderitem_count,
    }

    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.append({"A": "Product Id:", "B": product.id})
        ws.append({"A": "Product Manufacturer:", "B": product.manufacturer})
        ws.append({"A": "Product Model:", "B": product.model})
        ws.append({"A": "Product Price:", "B": product.price})
        ws.append({"A": "Product Quantity:", "B": product.quantity})
        ws.append({"A": "Limit per order:", "B": product.per_order_limit})
        ws.append({"A": "Warranty days:", "B": product.warranty_days})
        ws.append({"A": "Category:", "B": product.category.name})
        if product.image_url:
            ws.append({"A": "Image url:", "B": product.image_url})
        ws.append({"A": "Order count:", "B": product.order_count})
        ws.append({"A": "Order item count:", "B": product.orderitem_count})
        fp = BytesIO()
        setattr(fp, "name", f"product-{product.id}-report.xlsx")
        wb.save(fp)
        return Response(content=fp.getvalue(), media_type="application/vnd.ms-excel")

    return result


@router.get("/customers/{customer_id}")
async def customer_report(manager: AuthManagerDep, customer_id: int, fmt: Literal["json", "excel"]="json"):
    if not Permissions.check(manager, Permissions.READ_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient privileges!")

    customer = await Customer.get_or_none(id=customer_id).annotate(order_count=Count("orders"))

    result = {
        "id": customer.id,
        "first_name": customer.first_name,
        "last_name": customer.last_name,
        "email": customer.email,
        "phone_number": customer.phone_number,
        "order_count": customer.order_count,
        "orders": [],
        "returns": [],
        "total": 0,
        "averages": {}
    }

    orders = await Order.filter(customer=customer).all()
    for order in orders:
        result["orders"].append((await OrderPd.from_tortoise_orm(order)).model_dump())
        result["orders"][-1]["items"] = []

        order_total = 0
        orderitems = await order.items.all().select_related("orderitems")
        for prod in orderitems:
            result["orders"][-1]["items"].append(
                (await ProductPd.from_tortoise_orm(prod)).model_dump() |
                (await OrderItemPd.from_tortoise_orm(prod.orderitems)).model_dump(exclude={"id"})
            )
            order_total += prod.orderitems.quantity * prod.orderitems.price

        result["orders"][-1]["total"] = order_total
        result["total"] += order_total

        if order.creation_time > (datetime.now() - timedelta(days=365)).replace(tzinfo=UTC):
            date = f"{order.creation_time.month}.{order.creation_time.year}"
            if date not in result["averages"]:
                result["averages"][date] = 0

            if orderitems:
                result["averages"][date] += order_total / len(orderitems)

        for ret in await Return.filter(order=order).select_related("order_item", "order_item__product").all():
            result["returns"].append({
                "item": {
                    "model": ret.order_item.product.model,
                    "manufacturer": ret.order_item.product.manufacturer,
                    "price": ret.order_item.product.price,
                },
                "quantity": ret.quantity,
                "time": ret.creation_time,
            })

    if fmt == "excel":
        wb = Workbook()
        ws = CustomWorksheet(wb.active)
        ws.append({"A": "REPORT GENERATED AT", "B": datetime.now()})
        ws.append({})
        ws.append({"A": "Customer First Name", "B": customer.first_name})
        ws.append({"A": "Customer Last Name", "B": customer.last_name})
        ws.append({"A": "Customer Email", "B": customer.email})
        ws.append({"A": "Customer Phone Number", "B": str(customer.phone_number)})
        ws.append({"A": "Order count", "B": customer.order_count})
        ws.append({"A": "Total", "B": result["total"]})
        ws.append({})
        ws.append({"A": "Average per month:"})
        for month, money in result["averages"].items():
            ws.append({"A": datetime.strptime(month, "%m.%Y"), "B": money})
        ws.append({})
        ws.append({"A": upd_cell(ws("Orders:"), font=Font(bold=True), fill=mkfill("89EB34", .15))})
        f65 = lambda: mkfill("89EB34", .65)
        f85 = lambda: mkfill("89EB34", .85)
        for order in result["orders"]:
            ws.append({"A": upd_cell(ws("Status"), fill=f65()), "B": upd_cell(ws(order["status"]), fill=f65())})
            ws.append({"A": upd_cell(ws("Creation Time"), fill=f65()),
                       "B": upd_cell(ws(order["creation_time"].replace(tzinfo=None)), fill=f65())})
            ws.append({"A": upd_cell(ws("Address"), fill=f65()), "B": upd_cell(ws(order["address"]), fill=f65())})
            ws.append({"A": upd_cell(ws("Type"), fill=f65()), "B": upd_cell(ws(order["type"]), fill=f65())})
            ws.append({"A": upd_cell(ws("Total"), fill=f65()), "B": upd_cell(ws(order["total"]), fill=f65())})
            ws.append({"A": upd_cell(ws("Item name"), fill=f65()), "B": upd_cell(ws("Price"), fill=f65()),
                       "C": upd_cell(ws("Quantity"), fill=f65())})

            for item in order["items"]:
                ws.append({"A": upd_cell(ws(f"{item['manufacturer']} {item['model']}"), fill=f85()),
                           "B": upd_cell(ws(item["price"]), fill=f85()),
                           "C": upd_cell(ws(item["quantity"]), fill=f85())})
            ws.append({})

        ws.append({"A": upd_cell(ws("Returns:"), font=Font(bold=True), fill=mkfill("89EB34", .15))})
        ws.append({"A": upd_cell(ws("Item name"), fill=f65()), "B": upd_cell(ws("Quantity"), fill=f65()),
                   "C": upd_cell(ws("Time"), fill=f65()), "D": upd_cell(ws("Price (per item)"), fill=f65())})
        for ret in result["returns"]:
            ws.append({"A": upd_cell(ws(f"{ret['item']['manufacturer']} {ret['item']['model']}"), fill=f85()),
                       "B": upd_cell(ws(ret["quantity"]), fill=f85()),
                       "C": upd_cell(ws(ret["time"].replace(tzinfo=None)), fill=f85()),
                       "D": upd_cell(ws(ret["item"]["price"]), fill=f85())})

        ws.fix_formatting()
        fp = BytesIO()
        setattr(fp, "name", f"customer-{customer.id}-report.xlsx")
        wb.save(fp)
        return Response(content=fp.getvalue(), media_type="application/vnd.ms-excel")

    return result


@router.get("/categories/{category_id}")
async def category_report(manager: AuthManagerDep, category_id: int, fmt: Literal["json", "excel"]="json"):
    if not Permissions.check(manager, Permissions.READ_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient privileges!")

    category = await Category.get_or_none(id=category_id)

    result = {
        "id": category.id,
        "name": category.name,
        "description": category.description,
        "product_count": 0,
        "order_count": 0,
        "product_quantity": 0,
        "product_average_price": 0,
        "products": [],
    }

    products = await Product.filter(category=category).all()
    result["product_count"] = len(products)
    for product in products:
        result["product_quantity"] += product.quantity
        result["product_average_price"] += product.price
        result["products"].append((await ProductPd.from_tortoise_orm(product)).model_dump())

    result["order_count"] = await Order.filter(orderitems__product__id__in=[prod.id for prod in products]).count()
    result["product_average_price"] /= result["product_count"]

    if fmt == "excel":
        wb = Workbook()
        ws = CustomWorksheet(wb.active)
        ws.append({"A": "REPORT GENERATED AT", "B": datetime.now()})
        ws.append({})
        ws.append({"A": "Category Id:", "B": category_id})
        ws.append({"A": "Category Name:", "B": category.name})
        ws.append({"A": "Category Description:", "B": category.description})
        ws.append({})
        ws.append({"A": "Product count:", "B": result["product_count"]})
        ws.append({"A": "Order count:", "B": result["order_count"]})
        ws.append({"A": "Product quantity:", "B": result["product_quantity"]})
        ws.append({"A": "Product average price:", "B": result["product_average_price"]})
        ws.append({})
        ws.append({"A": "Products:"})
        ws.append({})
        ws.append({"A": "Model", "B": "Manufacturer", "C": "Price", "D": "Quantity", "E": "Limit Per order",
                   "F": "Warranty Days"})
        for prod in result["products"]:
            ws.append({"A": prod["model"], "B": prod["manufacturer"], "C": prod["price"], "D": prod["quantity"],
                       "E": prod["per_order_limit"], "F": prod["warranty_days"]})

        ws.fix_formatting()
        fp = BytesIO()
        setattr(fp, "name", f"category-{category.id}-report.xlsx")
        wb.save(fp)
        return Response(content=fp.getvalue(), media_type="application/vnd.ms-excel")

    return result
